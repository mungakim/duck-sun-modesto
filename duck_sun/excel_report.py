"""
Excel Report Generator for Duck Sun Modesto
Generates Excel (.xlsx) reports matching the PDF format - CENTERED LAYOUT

Weights: Google(6x), Accu(4x), Weather.com(4x), WUnderground(4x), NOAA(3x), Met.no(3x), OM(1x)
"""

import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Any
from zoneinfo import ZoneInfo

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Border, Side, Alignment,
        NamedStyle
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    Workbook = None

logger = logging.getLogger(__name__)

# Column offset for centering (start content at column C)
COL_OFFSET = 2


def col(n):
    """Get column letter with offset for centering."""
    return get_column_letter(n + COL_OFFSET)


def calculate_daily_stats_from_hourly(hourly_data: List[Dict], timezone: str = "America/Los_Angeles") -> Dict:
    """Calculate daily high/low from hourly data using meteorological day (6am-6am)."""
    daily_stats = {}
    tz = ZoneInfo(timezone)

    for record in hourly_data:
        try:
            t = record.get('time', '')
            val = record.get('temp_c', record.get('temperature_c'))
            if val is None:
                continue

            if '+' in t or 'Z' in t:
                dt = datetime.fromisoformat(t.replace('Z', '+00:00')).astimezone(tz)
            else:
                dt = datetime.fromisoformat(t)

            if dt.hour < 6:
                met_day = dt - timedelta(days=1)
            else:
                met_day = dt

            k = met_day.strftime('%Y-%m-%d')
            if k not in daily_stats:
                daily_stats[k] = {'temps': [], 'day_name': met_day.strftime('%a')}
            daily_stats[k]['temps'].append(float(val))

        except Exception as e:
            logger.debug(f"[calculate_daily_stats] Failed to parse record: {e}")
            continue

    result = {}
    for k, d in daily_stats.items():
        if d['temps']:
            result[k] = {
                'date': k,
                'day_name': d['day_name'],
                'high_f': round(max(d['temps']) * 1.8 + 32),
                'low_f': round(min(d['temps']) * 1.8 + 32)
            }

    return result


def calculate_weighted_average(values: List[Optional[float]], weights: List[float]) -> Optional[int]:
    """Calculate weighted average from values with weights."""
    total_val, total_weight = 0.0, 0.0

    for val, weight in zip(values, weights):
        if val is not None:
            total_val += val * weight
            total_weight += weight

    if total_weight > 0:
        return round(total_val / total_weight)
    return None


def calculate_weighted_average_excluding_om_max(
    values: List[Optional[float]],
    weights: List[float]
) -> tuple[Optional[int], set[int]]:
    """
    Calculate weighted average, excluding Open-Meteo (index 0) only if it has the max value.
    """
    valid_pairs = [(i, v) for i, v in enumerate(values) if v is not None]

    if not valid_pairs:
        return None, set()

    max_val = max(v for _, v in valid_pairs)

    om_val = values[0]
    excluded_indices = set()
    if om_val is not None and om_val == max_val:
        excluded_indices = {0}

    total_val, total_weight = 0.0, 0.0
    for i, v in valid_pairs:
        if i not in excluded_indices:
            total_val += v * weights[i]
            total_weight += weights[i]

    if total_weight > 0:
        return round(total_val / total_weight), excluded_indices

    if om_val is not None:
        return round(om_val), excluded_indices
    return None, excluded_indices


def calculate_clear_sky_ghi(hour: int, day_of_year: int, lat: float = 37.6391) -> float:
    """Calculate theoretical clear-sky Global Horizontal Irradiance (GHI)."""
    import math

    declination = 23.45 * math.sin(math.radians(360 * (284 + day_of_year) / 365))
    hour_angle = 15 * (hour - 12.5)

    lat_rad = math.radians(lat)
    decl_rad = math.radians(declination)
    hour_rad = math.radians(hour_angle)

    sin_elevation = (math.sin(lat_rad) * math.sin(decl_rad) +
                     math.cos(lat_rad) * math.cos(decl_rad) * math.cos(hour_rad))

    if sin_elevation <= 0:
        return 0.0

    seasonal_factor = 0.7 + 0.3 * math.cos(math.radians((day_of_year - 172) * 360 / 365))
    max_ghi = 900 * seasonal_factor

    ghi = max_ghi * sin_elevation

    return min(ghi, 900)


def estimate_irradiance_from_cloud_cover(cloud_cover: int, hour: int, day_of_year: int) -> float:
    """Estimate solar irradiance from cloud cover percentage."""
    clear_sky_ghi = calculate_clear_sky_ghi(hour, day_of_year)

    if clear_sky_ghi <= 0:
        return 0.0

    cloud_fraction = cloud_cover / 100.0
    attenuation = 1.0 - (0.7 * cloud_fraction)

    return round(clear_sky_ghi * attenuation, 1)


def get_solar_color_and_desc(risk_level: str, solar_value: float, condition: str = None) -> tuple:
    """Get cell color AND short description based on solar conditions.

    Descriptions are kept SHORT to fit in narrow Excel columns (max ~8 chars).
    """
    risk_upper = risk_level.upper()

    if 'TULE FOG' in risk_upper:
        return "B4A0C8", "Tule Fog"

    if 'CRITICAL' in risk_upper or 'ACTIVE FOG' in risk_upper:
        return "FFB4B4", "Fog"
    elif 'HIGH' in risk_upper or 'STRATUS' in risk_upper:
        return "FFD2A0", "Overcast"
    elif 'MODERATE' in risk_upper:
        return "FFFFB4", "Fog?"

    if condition and condition not in ('Unknown', 'Open-Meteo'):
        cond_lower = condition.lower()
        if 'rain' in cond_lower or 'storm' in cond_lower or 'shower' in cond_lower:
            desc = "Lt rain" if 'light' in cond_lower else "Rain" if 'rain' in cond_lower else "Storms"
            return "FFD2A0", desc
        elif 'fog' in cond_lower or 'mist' in cond_lower:
            return "FFFFB4", "Fog?"
        elif 'cloudy' in cond_lower:
            if 'partly' in cond_lower:
                return "C8E6FF", "P cloudy"
            elif 'mostly' in cond_lower:
                if solar_value < 50:
                    return "DCDCDC", "M cloudy"
                else:
                    return "C8E6FF", "M cloudy"
            else:
                return "DCDCDC", "Cloudy"
        elif 'clear' in cond_lower or 'sunny' in cond_lower:
            if solar_value >= 400:
                return "90EE90", "Sunny"
            elif solar_value >= 150:
                return "C8FFC8", "Sunny"
            else:
                return "C8E6FF", "Sunny"

    if solar_value < 50:
        return "DCDCDC", "Cloudy"
    elif solar_value < 150:
        return "C8E6FF", "Some Sun"
    elif solar_value < 400:
        return "C8FFC8", "Good Sun"
    else:
        return "90EE90", "Full Sun"


def get_daily_condition_display(condition: str, dewpoint_c: float = None, temp_c: float = None,
                                 visibility_low: bool = False) -> tuple:
    """Map condition to display text and color for daily descriptor."""
    if not condition or condition == "Unknown":
        return ("--", "F0F0F0")

    cond_lower = condition.lower()

    is_potential_fog = False
    if dewpoint_c is not None and temp_c is not None:
        temp_dewpoint_spread = temp_c - dewpoint_c
        if temp_dewpoint_spread < 2.0 and ('fog' in cond_lower or 'mist' in cond_lower):
            is_potential_fog = True

    if 'fog' in cond_lower or 'mist' in cond_lower:
        if is_potential_fog or visibility_low:
            return ("TULE FOG", "B40000")
        else:
            return ("Fog", "FFE6B4")

    if 'thunderstorm' in cond_lower or 'storm' in cond_lower:
        return ("Storms", "6464B4")
    elif 'heavy rain' in cond_lower:
        return ("Heavy Rain", "648CC8")
    elif 'rain shower' in cond_lower or 'showers' in cond_lower:
        return ("Showers", "8CAADC")
    elif 'light rain' in cond_lower:
        return ("Light Rain", "B4C8E6")
    elif 'drizzle' in cond_lower:
        return ("Drizzle", "B4C8E6")
    elif 'rain' in cond_lower:
        return ("Rain", "78A0D2")

    if 'snow' in cond_lower or 'sleet' in cond_lower or 'ice' in cond_lower:
        return ("SNOW", "C8DCFF")

    if 'overcast' in cond_lower:
        return ("Overcast", "C8C8C8")
    elif 'cloudy' in cond_lower:
        if 'partly' in cond_lower:
            return ("Partly Cloudy", "E6F5FF")
        elif 'mostly' in cond_lower:
            return ("Mostly Cloudy", "D2DCE6")
        else:
            return ("Cloudy", "C8D2DC")

    if 'clear' in cond_lower or 'sunny' in cond_lower:
        return ("Sunny", "FFFAC8")
    elif 'fair' in cond_lower:
        return ("Fair", "FAFADC")

    if 'haze' in cond_lower or 'smoke' in cond_lower:
        return ("SMOKE/HAZE", "FFC896")

    if 'wind' in cond_lower:
        return ("Windy", "E6F0FF")

    display = condition[:12] if len(condition) > 12 else condition
    return (display, "F5F5F5")


def generate_excel_report(
    om_data: Dict,
    noaa_data: Optional[List],
    met_data: Optional[List],
    accu_data: Optional[List],
    google_data: Optional[Dict] = None,
    weather_com_data: Optional[List] = None,
    wunderground_data: Optional[List] = None,
    df_analyzed: Any = None,
    output_path: Optional[Path] = None,
    mid_data: Optional[Dict] = None,
    precip_data: Optional[Dict] = None,
    noaa_daily_periods: Optional[Dict] = None,
    report_timestamp: Optional[datetime] = None
) -> Optional[Path]:
    """
    Generate Excel report with 7-source temperature grid and weighted consensus.
    CENTERED layout matching PDF format.
    """
    if not HAS_OPENPYXL:
        logger.error("[generate_excel_report] openpyxl not installed")
        return None

    logger.info("[generate_excel_report] Starting Excel generation (CENTERED)...")

    # Process data sources
    om_daily = om_data.get('daily_forecast', [])[:8]
    met_daily = calculate_daily_stats_from_hourly(met_data) if met_data else {}

    if noaa_daily_periods:
        noaa_daily = noaa_daily_periods
    else:
        noaa_daily = calculate_daily_stats_from_hourly(noaa_data) if noaa_data else {}

    # Process AccuWeather data
    accu_daily = {}
    if accu_data:
        for d in accu_data:
            if 'high_f' in d and 'low_f' in d:
                accu_daily[d['date']] = {
                    'high_f': int(d['high_f']),
                    'low_f': int(d['low_f'])
                }
            elif 'high_c' in d and 'low_c' in d:
                accu_daily[d['date']] = {
                    'high_f': round(d['high_c'] * 1.8 + 32),
                    'low_f': round(d['low_c'] * 1.8 + 32)
                }
        logger.info(f"[generate_excel_report] AccuWeather processed: {len(accu_daily)} days")

    # Process Google Weather data
    google_daily = {}
    if google_data:
        daily_list = google_data.get('daily', [])
        for d in daily_list:
            if 'high_f' in d and 'low_f' in d:
                google_daily[d['date']] = {
                    'high_f': int(d['high_f']),
                    'low_f': int(d['low_f'])
                }
            elif 'high_c' in d and 'low_c' in d:
                google_daily[d['date']] = {
                    'high_f': round(d['high_c'] * 1.8 + 32),
                    'low_f': round(d['low_c'] * 1.8 + 32)
                }
        logger.info(f"[generate_excel_report] Google Weather processed: {len(google_daily)} days")

    # Process Weather.com data
    weather_com_daily = {}
    if weather_com_data:
        for d in weather_com_data:
            if 'high_f' in d and 'low_f' in d:
                weather_com_daily[d['date']] = {
                    'high_f': int(d['high_f']),
                    'low_f': int(d['low_f'])
                }
        logger.info(f"[generate_excel_report] Weather.com processed: {len(weather_com_daily)} days")

    # Process Weather Underground data
    wunderground_daily = {}
    if wunderground_data:
        for d in wunderground_data:
            if 'high_f' in d and 'low_f' in d:
                wunderground_daily[d['date']] = {
                    'high_f': int(d['high_f']),
                    'low_f': int(d['low_f'])
                }
        logger.info(f"[generate_excel_report] Weather Underground processed: {len(wunderground_daily)} days")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Forecast"

    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Use timestamp
    if report_timestamp:
        report_time = report_timestamp
    else:
        report_time = datetime.now(ZoneInfo("America/Los_Angeles"))
    timestamp_str = report_time.strftime("%A, %B %d, %Y %H:%M:%S")

    # Set column widths - balanced to fit on one landscape page with readable text
    ws.column_dimensions['A'].width = 1   # Left margin spacer
    ws.column_dimensions['B'].width = 1   # Left margin spacer
    ws.column_dimensions[col(1)].width = 11  # PGE CITYGATE / MID GAS NOM label column (C) - wider for labels
    # Solar forecast columns (9AM-4PM) = col(2) through col(9) - UNIFORM width
    for i in range(2, 10):
        ws.column_dimensions[col(i)].width = 8  # Uniform width for solar columns
    # Remaining data columns
    for i in range(10, 20):
        ws.column_dimensions[col(i)].width = 7.5

    # Page setup: ZERO top margin to move report up, LANDSCAPE, fit to ONE page
    ws.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0,
        bottom=0.5,
        header=0,
        footer=0.1
    )
    # Force landscape orientation
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 1  # 1 = Letter size in openpyxl
    # Fit to exactly one page
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False  # Disabled to push content to TOP of page

    # =====================
    # ROW 1: Title (CENTERED across full width) - moved up from row 2
    # =====================
    ws.merge_cells(f'{col(1)}1:{col(18)}1')
    title_cell = ws[f'{col(1)}1']
    title_cell.value = "MODESTO, CA - DAILY WEATHER FORECAST"
    title_cell.font = Font(name='Arial', size=14, bold=True, color='003C78')
    title_cell.alignment = center_align

    # =====================
    # ROW 2: Timestamp (CENTERED) - moved up from row 3
    # =====================
    ws.merge_cells(f'{col(1)}2:{col(18)}2')
    ts_cell = ws[f'{col(1)}2']
    ts_cell.value = timestamp_str
    ts_cell.font = Font(name='Arial', size=11, bold=True)
    ts_cell.alignment = center_align

    # =====================
    # ROW 3: PGE CITYGATE - Label cell (NO border) + MERGED input cell (D+E) WITH border
    # =====================
    # Create explicit border for PGE CITYGATE input cell only
    pge_input_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    pge_label = ws[f'{col(1)}3']
    pge_label.value = "PGE CITYGATE:   "
    pge_label.font = Font(name='Arial', size=8, bold=True)
    pge_label.alignment = center_align
    # NO border on the label cell
    # Input cell for PGE CITYGATE - MERGE col(2) and col(3) for wider cell
    ws.merge_cells(f'{col(2)}3:{col(3)}3')
    pge_input = ws[f'{col(2)}3']
    pge_input.border = pge_input_border  # Border only on input cell
    pge_input.alignment = center_align
    pge_input.font = Font(name='Arial', size=8, bold=True, color='9E470E')  # Bold, Orange darker
    pge_input.number_format = '"$"0.000'  # Dollar sign with exactly 3 decimal places
    # Apply border to merged range
    ws[f'{col(3)}3'].border = pge_input_border

    # =====================
    # ROW 5-8: MID GAS NOM - Left column dates (MM/DD/YY), Right column MERGED (D+E)
    # =====================
    ws[f'{col(1)}5'] = "MID GAS NOM:"
    ws[f'{col(1)}5'].font = Font(name='Arial', size=8, bold=True)
    ws[f'{col(1)}5'].alignment = left_align

    # Create 3 rows of cells with merged right column
    for row_idx in range(6, 9):
        # Left column - date cells with MM/DD/YY format
        cell_left = ws[f'{col(1)}{row_idx}']
        cell_left.border = thick_border
        cell_left.alignment = center_align
        cell_left.font = Font(name='Arial', size=8, bold=True, color='9E470E')  # Bold, Orange darker
        cell_left.number_format = 'MM/DD/YY'  # Date format: 11/09/26

        # Right column - MERGE col(2) and col(3) for wider cell
        ws.merge_cells(f'{col(2)}{row_idx}:{col(3)}{row_idx}')
        cell_right = ws[f'{col(2)}{row_idx}']
        cell_right.border = thick_border
        cell_right.alignment = center_align
        cell_right.font = Font(name='Arial', size=8, bold=True, color='9E470E')  # Bold, Orange darker
        cell_right.number_format = '#,##0" MMBtus"'  # Format: 8,000 MMBtus
        # Apply border to merged range
        ws[f'{col(3)}{row_idx}'].border = thick_border

    # =====================
    # MID WEATHER 48-HOUR SUMMARY (right side) - Extended range to fit title
    # =====================
    ws.merge_cells(f'{col(11)}4:{col(17)}4')
    mid_header = ws[f'{col(11)}4']
    mid_header.value = "MID WEATHER 48-HOUR SUMMARY"
    mid_header.font = Font(name='Arial', size=9, bold=True, color='003C78')
    mid_header.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    mid_header.alignment = center_align
    mid_header.border = thin_border
    # Apply border to all cells in merged range
    for c in range(11, 18):
        ws[f'{col(c)}4'].border = thin_border

    # Headers: High, Low, Rain (no empty cell before High)
    for c, label in [(14, 'High'), (15, 'Low'), (16, 'Rain')]:
        cell = ws[f'{col(c)}5']
        cell.value = label
        cell.font = Font(name='Arial', size=7, bold=True)
        cell.alignment = center_align
        cell.border = thin_border

    if mid_data:
        today_data = mid_data.get('today', {})
        yest_data = mid_data.get('yesterday', {})

        # TODAY row
        ws.merge_cells(f'{col(12)}6:{col(13)}6')
        ws[f'{col(12)}6'] = "TODAY"
        ws[f'{col(12)}6'].font = Font(name='Arial', size=7, bold=True)
        ws[f'{col(12)}6'].alignment = center_align
        ws[f'{col(12)}6'].border = thin_border
        ws[f'{col(13)}6'].border = thin_border

        ws[f'{col(14)}6'] = f"{today_data.get('high', '--')}F"
        ws[f'{col(14)}6'].fill = PatternFill(start_color="FFC8B4", end_color="FFC8B4", fill_type="solid")
        ws[f'{col(14)}6'].alignment = center_align
        ws[f'{col(14)}6'].border = thin_border

        ws[f'{col(15)}6'] = f"{today_data.get('low', '--')}F"
        ws[f'{col(15)}6'].fill = PatternFill(start_color="B4D2FF", end_color="B4D2FF", fill_type="solid")
        ws[f'{col(15)}6'].alignment = center_align
        ws[f'{col(15)}6'].border = thin_border

        ws[f'{col(16)}6'] = f"{today_data.get('rain', '0.00')}\""
        ws[f'{col(16)}6'].alignment = center_align
        ws[f'{col(16)}6'].border = thin_border

        # YESTERDAY row
        ws.merge_cells(f'{col(12)}7:{col(13)}7')
        ws[f'{col(12)}7'] = "YESTERDAY"
        ws[f'{col(12)}7'].font = Font(name='Arial', size=7, bold=True)
        ws[f'{col(12)}7'].alignment = center_align
        ws[f'{col(12)}7'].border = thin_border
        ws[f'{col(13)}7'].border = thin_border

        ws[f'{col(14)}7'] = f"{yest_data.get('high', '--')}F"
        ws[f'{col(14)}7'].fill = PatternFill(start_color="FFC8B4", end_color="FFC8B4", fill_type="solid")
        ws[f'{col(14)}7'].alignment = center_align
        ws[f'{col(14)}7'].border = thin_border

        ws[f'{col(15)}7'] = f"{yest_data.get('low', '--')}F"
        ws[f'{col(15)}7'].fill = PatternFill(start_color="B4D2FF", end_color="B4D2FF", fill_type="solid")
        ws[f'{col(15)}7'].alignment = center_align
        ws[f'{col(15)}7'].border = thin_border

        ws[f'{col(16)}7'] = f"{yest_data.get('rain', '0.00')}\""
        ws[f'{col(16)}7'].alignment = center_align
        ws[f'{col(16)}7'].border = thin_border

        # Records row
        if 'record_high_temp' in mid_data:
            ws.merge_cells(f'{col(12)}8:{col(16)}8')
            rec_cell = ws[f'{col(12)}8']
            rec_hi = mid_data.get('record_high_temp', '--')
            rec_hi_yr = mid_data.get('record_high_year', '')
            rec_lo = mid_data.get('record_low_temp', '--')
            rec_lo_yr = mid_data.get('record_low_year', '')
            rec_cell.value = f"Records: Hi {rec_hi}F({rec_hi_yr}) Lo {rec_lo}F({rec_lo_yr})"
            rec_cell.font = Font(name='Arial', size=6, italic=True)
            rec_cell.alignment = center_align
            # Apply border to all cells in merged range
            for c in range(12, 17):
                ws[f'{col(c)}8'].border = thin_border

    # =====================
    # TEMPERATURE GRID starting at Row 10 (moved up from row 11)
    # =====================
    SOURCE_WEIGHT_DISPLAY = {
        'OPEN-METEO': '1.0',
        'NOAA (GOV)': '3.0',
        'MET.NO (EU)': '3.0',
        'ACCUWEATHER': '4.0',
        'WEATHER.COM': '4.0',
        'WUNDERGRND': '4.0',
        'GOOGLE (AI)': '6.0',
    }

    # URLs for clickable source links
    SOURCE_URLS = {
        'NOAA (GOV)': 'https://forecast.weather.gov/MapClick.php?lat=37.6684&lon=-120.99',
        'ACCUWEATHER': 'https://www.accuweather.com/en/us/modesto/95354/daily-weather-forecast/327145?page=0',
        'WEATHER.COM': 'https://weather.com/weather/tenday/l/USCA0714',
        'WUNDERGRND': 'https://www.wunderground.com/forecast/us/ca/modesto/95350?cm_ven=localwx_10day',
    }

    DAY_COLORS = [
        "FFF0F0", "F0FFF0", "F0F8FF", "FFFFF0",
        "FFF5EE", "F5FFFA", "F8F8FF", "FFFAF0",
    ]

    # Build merged conditions map
    daily_conditions = {}
    for day_record in om_daily:
        date_key = day_record.get('date', '')
        condition = day_record.get('condition', 'Unknown')
        if condition and condition != 'Unknown':
            daily_conditions[date_key] = {'condition': condition, 'source': 'Open-Meteo'}

    if accu_data:
        for day_record in accu_data:
            date_key = day_record.get('date', '')
            condition = day_record.get('condition', '')
            if condition and condition != 'Unknown':
                daily_conditions[date_key] = {'condition': condition, 'source': 'AccuWeather'}

    if google_data:
        for day_record in google_data.get('daily', []):
            date_key = day_record.get('date', '')
            condition = day_record.get('condition', '')
            if condition and condition != 'Unknown':
                daily_conditions[date_key] = {'condition': condition, 'source': 'Google'}

    # Row 10: Condition descriptors - NO borders on merged empty cells
    grid_row = 10
    ws.merge_cells(f'{col(1)}{grid_row}:{col(2)}{grid_row}')
    ws[f'{col(1)}{grid_row}'] = ""
    # Merged col(1)+col(2) - no borders for this empty area

    for i, day in enumerate(om_daily):
        date_key = day.get('date', '')
        condition_info = daily_conditions.get(date_key, {'condition': 'Unknown'})
        condition = condition_info['condition']
        display_text, bg_color = get_daily_condition_display(condition)

        col_hi = col(3 + i * 2)
        col_lo = col(4 + i * 2)

        ws.merge_cells(f'{col_hi}{grid_row}:{col_lo}{grid_row}')
        cell = ws[f'{col_hi}{grid_row}']
        cell.value = display_text
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.font = Font(name='Arial', size=7)
        cell.alignment = center_align
        cell.border = thin_border
        # Apply border to second cell of merged range
        ws[f'{col_lo}{grid_row}'].border = thin_border

    # Row 11: Day names header - empty merged cell for label area
    grid_row = 11
    ws.merge_cells(f'{col(1)}{grid_row}:{col(2)}{grid_row}')
    empty_header = ws[f'{col(1)}{grid_row}']
    empty_header.value = ""
    empty_header.fill = PatternFill(start_color="003C78", end_color="003C78", fill_type="solid")
    empty_header.border = thin_border
    ws[f'{col(2)}{grid_row}'].border = thin_border

    for i, day in enumerate(om_daily):
        label = "TODAY" if i == 0 else day.get('day_name', '')[:3].upper()
        col_hi = col(3 + i * 2)
        col_lo = col(4 + i * 2)

        ws.merge_cells(f'{col_hi}{grid_row}:{col_lo}{grid_row}')
        cell = ws[f'{col_hi}{grid_row}']
        cell.value = label
        cell.fill = PatternFill(start_color="003C78", end_color="003C78", fill_type="solid")
        cell.font = Font(name='Arial', size=8, bold=True, color='FFFFFF')
        cell.alignment = center_align
        cell.border = thin_border
        # Apply border to second cell of merged range
        ws[f'{col_lo}{grid_row}'].border = thin_border

    # Row 12: Dates with SOURCE label (instead of DATE)
    grid_row = 12
    ws.merge_cells(f'{col(1)}{grid_row}:{col(2)}{grid_row}')
    source_header = ws[f'{col(1)}{grid_row}']
    source_header.value = "SOURCE"
    source_header.fill = PatternFill(start_color="466EA0", end_color="466EA0", fill_type="solid")
    source_header.font = Font(name='Arial', size=7, bold=True, color='FFFFFF')
    source_header.alignment = center_align
    source_header.border = thin_border
    ws[f'{col(2)}{grid_row}'].border = thin_border

    for i, day in enumerate(om_daily):
        date_str = day.get('date', '')[5:]
        col_hi = col(3 + i * 2)
        col_lo = col(4 + i * 2)

        ws.merge_cells(f'{col_hi}{grid_row}:{col_lo}{grid_row}')
        cell = ws[f'{col_hi}{grid_row}']
        cell.value = date_str
        cell.fill = PatternFill(start_color="466EA0", end_color="466EA0", fill_type="solid")
        cell.font = Font(name='Arial', size=7, color='FFFFFF')
        cell.alignment = center_align
        cell.border = thin_border
        # Apply border to second cell of merged range
        ws[f'{col_lo}{grid_row}'].border = thin_border

    # Pre-calculate excluded highs
    weights = [1.0, 3.0, 3.0, 4.0, 4.0, 4.0, 6.0]
    excluded_highs = {}
    for i, day in enumerate(om_daily):
        k = day.get('date', '')
        hi_vals = [
            day.get('high_f'),
            noaa_daily.get(k, {}).get('high_f'),
            met_daily.get(k, {}).get('high_f'),
            accu_daily.get(k, {}).get('high_f'),
            weather_com_daily.get(k, {}).get('high_f'),
            wunderground_daily.get(k, {}).get('high_f'),
            google_daily.get(k, {}).get('high_f')
        ]
        valid_highs = [(idx, v) for idx, v in enumerate(hi_vals) if v is not None]
        if valid_highs:
            max_high = max(v for _, v in valid_highs)
            om_high = hi_vals[0]
            if om_high is not None and om_high == max_high:
                excluded_highs[i] = {0}
            else:
                excluded_highs[i] = set()
        else:
            excluded_highs[i] = set()

    # Source rows
    sources = [
        ('OPEN-METEO', lambda d, k: (d.get('high_f'), d.get('low_f')), 0),
        ('NOAA (GOV)', lambda d, k: (noaa_daily.get(k, {}).get('high_f'), noaa_daily.get(k, {}).get('low_f')), 1),
        ('MET.NO (EU)', lambda d, k: (met_daily.get(k, {}).get('high_f'), met_daily.get(k, {}).get('low_f')), 2),
        ('ACCUWEATHER', lambda d, k: (accu_daily.get(k, {}).get('high_f'), accu_daily.get(k, {}).get('low_f')), 3),
        ('WEATHER.COM', lambda d, k: (weather_com_daily.get(k, {}).get('high_f'), weather_com_daily.get(k, {}).get('low_f')), 4),
        ('WUNDERGRND', lambda d, k: (wunderground_daily.get(k, {}).get('high_f'), wunderground_daily.get(k, {}).get('low_f')), 5),
        ('GOOGLE (AI)', lambda d, k: (google_daily.get(k, {}).get('high_f'), google_daily.get(k, {}).get('low_f')), 6),
    ]

    for src_idx, (label, getter, source_index) in enumerate(sources):
        grid_row = 13 + src_idx

        # MERGED col(1)+col(2) for wider source label (no weight shown, centered)
        ws.merge_cells(f'{col(1)}{grid_row}:{col(2)}{grid_row}')
        source_cell = ws[f'{col(1)}{grid_row}']
        source_cell.value = label  # Just the source name, no weight
        source_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        source_cell.alignment = center_align  # Centered
        source_cell.border = thin_border
        ws[f'{col(2)}{grid_row}'].border = thin_border

        # Add clickable hyperlink if URL exists for this source
        if label in SOURCE_URLS:
            source_cell.hyperlink = SOURCE_URLS[label]
            source_cell.font = Font(name='Arial', size=7, bold=True, color='0000FF', underline='single')
        else:
            source_cell.font = Font(name='Arial', size=7, bold=True)

        for i, day in enumerate(om_daily):
            k = day.get('date', '')
            v1, v2 = getter(day, k)
            is_excluded_high = source_index in excluded_highs.get(i, set())

            day_color = DAY_COLORS[i % len(DAY_COLORS)]

            col_hi = col(3 + i * 2)
            col_lo = col(4 + i * 2)

            cell_hi = ws[f'{col_hi}{grid_row}']
            if is_excluded_high and v1 is not None:
                cell_hi.value = "-"
            else:
                cell_hi.value = str(v1) if v1 else "--"
            cell_hi.fill = PatternFill(start_color=day_color, end_color=day_color, fill_type="solid")
            cell_hi.font = Font(name='Arial', size=9)
            cell_hi.alignment = center_align
            cell_hi.border = thin_border

            cell_lo = ws[f'{col_lo}{grid_row}']
            cell_lo.value = str(v2) if v2 else "--"
            cell_lo.fill = PatternFill(start_color=day_color, end_color=day_color, fill_type="solid")
            cell_lo.font = Font(name='Arial', size=9)
            cell_lo.alignment = center_align
            cell_lo.border = thin_border

    # Weighted Averages row - MERGED col(1)+col(2) for wider label
    grid_row = 20
    ws.merge_cells(f'{col(1)}{grid_row}:{col(2)}{grid_row}')
    wtd_cell = ws[f'{col(1)}{grid_row}']
    wtd_cell.value = "Wtd. Average"
    wtd_cell.fill = PatternFill(start_color="FFDC64", end_color="FFDC64", fill_type="solid")
    wtd_cell.font = Font(name='Arial', size=7, bold=True)
    wtd_cell.alignment = center_align
    wtd_cell.border = thin_border
    ws[f'{col(2)}{grid_row}'].border = thin_border

    for i, day in enumerate(om_daily):
        k = day.get('date', '')
        hi_vals = [
            day.get('high_f'),
            noaa_daily.get(k, {}).get('high_f'),
            met_daily.get(k, {}).get('high_f'),
            accu_daily.get(k, {}).get('high_f'),
            weather_com_daily.get(k, {}).get('high_f'),
            wunderground_daily.get(k, {}).get('high_f'),
            google_daily.get(k, {}).get('high_f')
        ]
        lo_vals = [
            day.get('low_f'),
            noaa_daily.get(k, {}).get('low_f'),
            met_daily.get(k, {}).get('low_f'),
            accu_daily.get(k, {}).get('low_f'),
            weather_com_daily.get(k, {}).get('low_f'),
            wunderground_daily.get(k, {}).get('low_f'),
            google_daily.get(k, {}).get('low_f')
        ]

        avg_hi, _ = calculate_weighted_average_excluding_om_max(hi_vals, weights)
        avg_lo = calculate_weighted_average(lo_vals, weights)

        col_hi = col(3 + i * 2)
        col_lo = col(4 + i * 2)

        cell_hi = ws[f'{col_hi}{grid_row}']
        cell_hi.value = str(avg_hi) if avg_hi else "--"
        cell_hi.fill = PatternFill(start_color="FFDC64", end_color="FFDC64", fill_type="solid")
        cell_hi.font = Font(name='Arial', size=9, bold=True)
        cell_hi.alignment = center_align
        cell_hi.border = thin_border

        cell_lo = ws[f'{col_lo}{grid_row}']
        cell_lo.value = str(avg_lo) if avg_lo else "--"
        cell_lo.fill = PatternFill(start_color="FFDC64", end_color="FFDC64", fill_type="solid")
        cell_lo.font = Font(name='Arial', size=9, bold=True)
        cell_lo.alignment = center_align
        cell_lo.border = thin_border

    # PRECIP % row - MERGED col(1)+col(2) for wider label
    grid_row = 21
    ws.merge_cells(f'{col(1)}{grid_row}:{col(2)}{grid_row}')
    precip_cell = ws[f'{col(1)}{grid_row}']
    precip_cell.value = "PRECIP %"
    precip_cell.fill = PatternFill(start_color="B4D2FF", end_color="B4D2FF", fill_type="solid")
    precip_cell.font = Font(name='Arial', size=7, bold=True)
    precip_cell.alignment = center_align
    precip_cell.border = thin_border
    ws[f'{col(2)}{grid_row}'].border = thin_border

    for i, day in enumerate(om_daily):
        k = day.get('date', '')
        precip_pct = 0
        if precip_data and k in precip_data:
            precip_pct = precip_data[k].get('consensus', 0)
        else:
            precip_pct = day.get('precip_prob', 0)

        if precip_pct >= 50:
            fill_color = "6496FF"
        elif precip_pct >= 25:
            fill_color = "B4D2FF"
        else:
            fill_color = DAY_COLORS[i % len(DAY_COLORS)]

        col_hi = col(3 + i * 2)
        col_lo = col(4 + i * 2)

        ws.merge_cells(f'{col_hi}{grid_row}:{col_lo}{grid_row}')
        cell = ws[f'{col_hi}{grid_row}']
        cell.value = f"{precip_pct}%"
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.font = Font(name='Arial', size=9)
        cell.alignment = center_align
        cell.border = thin_border

        # Apply border to merged cell's right side too
        cell_lo_ref = ws[f'{col_lo}{grid_row}']
        cell_lo_ref.border = thin_border

    # Precip source note
    grid_row = 22
    ws.merge_cells(f'{col(12)}{grid_row}:{col(18)}{grid_row}')
    note_cell = ws[f'{col(12)}{grid_row}']
    note_cell.value = "PRECIP = Google (0-72hr) > AccuWeather (72hr+) > Open-Meteo"
    note_cell.font = Font(name='Arial', size=6, italic=True, color='505050')
    note_cell.alignment = Alignment(horizontal='right', vertical='center')

    # =====================
    # SOLAR FORECAST GRID (also centered)
    # =====================
    grid_row = 24
    ws[f'{col(1)}{grid_row}'] = "SOLAR FORECAST (GOOGLE AI WEATHER API) - W/m² Irradiance"
    ws[f'{col(1)}{grid_row}'].font = Font(name='Arial', size=10, bold=True, color='003C78')

    tz = ZoneInfo("America/Los_Angeles")
    forecast_dates = [(datetime.now(tz) + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(0, 4)]

    # Build duck curve data
    duck_data = {d: [] for d in forecast_dates}
    google_hourly = google_data.get('hourly', []) if google_data else []

    for hour_record in google_hourly:
        try:
            time_str = hour_record.get('time', '')
            if not time_str:
                continue

            if 'Z' in time_str:
                dt = datetime.fromisoformat(time_str.replace('Z', '+00:00')).astimezone(tz)
            else:
                dt = datetime.fromisoformat(time_str).astimezone(tz)

            row_date = dt.strftime('%Y-%m-%d')
            row_hour = dt.hour

            if row_date in forecast_dates and 9 <= row_hour <= 16:
                cloud_cover = hour_record.get('cloud_cover', 50)
                day_of_year = dt.timetuple().tm_yday
                condition = hour_record.get('condition', 'Unknown')

                irradiance = estimate_irradiance_from_cloud_cover(cloud_cover, row_hour, day_of_year)

                condition_lower = condition.lower()
                if 'rain' in condition_lower or 'storm' in condition_lower:
                    risk = 'HIGH'
                elif cloud_cover >= 90:
                    risk = 'MODERATE'
                elif cloud_cover >= 70:
                    risk = 'LOW-MOD'
                else:
                    risk = 'LOW'

                duck_data[row_date].append({
                    'hour': row_hour,
                    'solar': irradiance,
                    'risk': risk,
                    'condition': condition
                })
        except Exception as e:
            logger.debug(f"[generate_excel_report] Error processing Google hour: {e}")
            continue

    # Fill gaps for today
    today = datetime.now(tz).strftime('%Y-%m-%d')
    if today in forecast_dates and df_analyzed is not None:
        existing_hours = {h['hour'] for h in duck_data.get(today, [])}
        missing_duck_hours = [h for h in range(9, 17) if h not in existing_hours]

        if missing_duck_hours:
            for _, row in df_analyzed.iterrows():
                try:
                    row_date = row['time'].strftime('%Y-%m-%d')
                    row_hour = row['time'].hour
                    if row_date == today and row_hour in missing_duck_hours:
                        solar_val = row.get('solar_adjusted', 0)
                        if solar_val == 0:
                            solar_val = row.get('solar_raw', 0)

                        duck_data[today].append({
                            'hour': row_hour,
                            'solar': solar_val,
                            'risk': row.get('risk_level', 'LOW'),
                            'condition': None
                        })
                except Exception:
                    continue

            duck_data[today].sort(key=lambda x: x['hour'])

    # Solar header row
    grid_row = 25
    header_labels = ['DATE', '9AM', '10', '11', '12PM', '1', '2', '3', '4PM']
    for col_idx, label in enumerate(header_labels):
        col_letter = col(1 + col_idx)
        cell = ws[f'{col_letter}{grid_row}']
        cell.value = label
        cell.fill = PatternFill(start_color="003C78", end_color="003C78", fill_type="solid")
        cell.font = Font(name='Arial', size=7, bold=True, color='FFFFFF')
        cell.alignment = center_align
        cell.border = thin_border

    # Solar data rows
    for date_idx, d in enumerate(forecast_dates):
        grid_row = 26 + date_idx * 2

        date_obj = datetime.strptime(d, '%Y-%m-%d')
        day_name = date_obj.strftime('%A')

        ws.merge_cells(f'{col(1)}{grid_row}:{col(1)}{grid_row + 1}')
        date_cell = ws[f'{col(1)}{grid_row}']
        date_cell.value = f"{d[5:]}\n{day_name}"
        date_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        date_cell.font = Font(name='Arial', size=7, bold=True)
        date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        date_cell.border = thin_border
        # Apply border to bottom cell of merged range
        ws[f'{col(1)}{grid_row + 1}'].border = thin_border

        hours_dict = {h['hour']: h for h in duck_data.get(d, [])}

        for h_idx in range(8):
            h_data = hours_dict.get(9 + h_idx, {'solar': 0, 'risk': 'LOW', 'condition': 'Unknown'})
            solar_display = h_data['solar'] * 1.15
            condition = h_data.get('condition', 'Unknown')

            color_hex, risk_desc = get_solar_color_and_desc(h_data['risk'], solar_display, condition)

            col_letter = col(2 + h_idx)

            val_cell = ws[f'{col_letter}{grid_row}']
            val_cell.value = int(solar_display)
            val_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            val_cell.font = Font(name='Arial', size=7)
            val_cell.alignment = center_align
            val_cell.border = thin_border

            desc_cell = ws[f'{col_letter}{grid_row + 1}']
            desc_cell.value = risk_desc
            desc_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            desc_cell.font = Font(name='Arial', size=6, italic=True)
            desc_cell.alignment = center_align
            desc_cell.border = thin_border

    # Legend row - row 35 with empty row 34 as gap after solar forecast
    grid_row = 35
    legend_items = [
        ("Cloudy", "DCDCDC"),
        ("Some Sun", "C8E6FF"),
        ("Good Sun", "C8FFC8"),
        ("Full Sun", "90EE90"),
        ("Fog Possible", "FFFFB4"),
        ("Heavy Clouds", "FFD2A0"),
        ("Dense Fog", "FFB4B4"),
        ("Tule Fog", "B4A0C8"),
    ]

    for col_idx, (label, color) in enumerate(legend_items):
        col_letter = col(2 + col_idx)  # Shifted right by 1 to align with values above
        cell = ws[f'{col_letter}{grid_row}']
        cell.value = label
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(name='Arial', size=6)
        cell.alignment = center_align
        cell.border = thin_border

    # Save workbook
    if output_path is None:
        pacific = ZoneInfo("America/Los_Angeles")
        now = datetime.now(pacific)
        timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
        output_path = Path("reports") / now.strftime("%Y-%m") / now.strftime("%Y-%m-%d") / f"daily_forecast_{timestamp}.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb.save(str(output_path))
        logger.info(f"[generate_excel_report] Excel saved to: {output_path}")
        return output_path
    except Exception as e:
        logger.error(f"[generate_excel_report] Failed to save Excel: {e}", exc_info=True)
        return None


if __name__ == "__main__":
    import asyncio
    from duck_sun.providers.open_meteo import fetch_open_meteo
    from dotenv import load_dotenv

    load_dotenv()
    logging.basicConfig(level=logging.INFO)

    async def test():
        print("=== Testing Excel Report Generator ===\n")
        om_data = await fetch_open_meteo(days=8)
        excel_path = generate_excel_report(
            om_data=om_data,
            noaa_data=None,
            met_data=None,
            accu_data=None
        )
        if excel_path:
            print(f"\n Excel generated: {excel_path}")

    asyncio.run(test())
