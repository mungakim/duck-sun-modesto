"""
Excel Report Generator for Duck Sun Modesto
Generates .xlsx files with the same data as the PDF report.

This module runs ALONGSIDE the PDF generator - both outputs are created.
"""

import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Any
from zoneinfo import ZoneInfo

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    Workbook = None

import pandas as pd

logger = logging.getLogger(__name__)

# Reuse calculation functions from pdf_report
from duck_sun.pdf_report import (
    calculate_daily_stats_from_hourly,
    calculate_weighted_average,
    calculate_weighted_average_excluding_om_max,
    calculate_clear_sky_ghi,
    estimate_irradiance_from_cloud_cover,
    get_solar_color_and_desc,
    get_daily_condition_display,
)


def rgb_to_hex(r: int, g: int, b: int) -> str:
    """Convert RGB values to hex color string for openpyxl."""
    return f"{r:02X}{g:02X}{b:02X}"


# Day column colors (pastels for readability) - matching PDF
DAY_COLORS = [
    (255, 240, 240),  # Day 0: Light pink
    (240, 255, 240),  # Day 1: Light green
    (240, 248, 255),  # Day 2: Light blue
    (255, 255, 240),  # Day 3: Light yellow
    (255, 245, 238),  # Day 4: Light peach
    (245, 255, 250),  # Day 5: Mint cream
    (248, 248, 255),  # Day 6: Ghost white
    (255, 250, 240),  # Day 7: Floral white
]

# Source weights for display (calibrated Jan 2026)
SOURCE_WEIGHT_DISPLAY = {
    'OPEN-METEO': '1.0',
    'NOAA (GOV)': '3.0',
    'MET.NO (EU)': '3.0',
    'ACCUWEATHER': '4.0',
    'WEATHER.COM': '4.0',
    'WUNDERGRND': '4.0',
    'GOOGLE (AI)': '6.0',
}

# URLs for clickable source links
SOURCE_URLS = {
    'NOAA (GOV)': 'https://forecast.weather.gov/MapClick.php?lat=37.6684&lon=-120.99',
    'ACCUWEATHER': 'https://www.accuweather.com/en/us/modesto/95354/daily-weather-forecast/327145?page=0',
    'WEATHER.COM': 'https://weather.com/weather/tenday/l/USCA0714',
    'WUNDERGRND': 'https://www.wunderground.com/forecast/us/ca/modesto/95350?cm_ven=localwx_10day',
}


def create_thin_border():
    """Create a thin border style for cells."""
    thin = Side(style='thin', color='808080')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def generate_xlsx_report(
    om_data: Dict,
    noaa_data: Optional[List],
    met_data: Optional[List],
    accu_data: Optional[List],
    google_data: Optional[Dict] = None,
    weather_com_data: Optional[List] = None,
    wunderground_data: Optional[List] = None,
    df_analyzed: pd.DataFrame = None,
    fog_critical_hours: int = 0,
    output_path: Optional[Path] = None,
    mid_data: Optional[Dict] = None,
    hrrr_data: Optional[Dict] = None,
    precip_data: Optional[Dict] = None,
    degraded_sources: Optional[List[str]] = None,
    noaa_daily_periods: Optional[Dict] = None,
    report_timestamp: Optional[datetime] = None
) -> Optional[Path]:
    """
    Generate Excel report with 7-source temperature grid and weighted consensus.

    This function accepts the SAME parameters as generate_pdf_report for 1:1 parity.
    """

    if not HAS_OPENPYXL:
        logger.error("[generate_xlsx_report] openpyxl not installed")
        return None

    logger.info("[generate_xlsx_report] Starting Excel generation...")

    # Process data sources (same as PDF)
    om_daily = om_data.get('daily_forecast', [])[:8]
    met_daily = calculate_daily_stats_from_hourly(met_data) if met_data else {}

    # PRIORITY: Use NOAA Period Data if available (matches website)
    if noaa_daily_periods:
        logger.info("[generate_xlsx_report] Using NOAA Period Data (Website Match)")
        noaa_daily = noaa_daily_periods
    else:
        logger.info("[generate_xlsx_report] Falling back to NOAA hourly aggregation")
        noaa_daily = calculate_daily_stats_from_hourly(noaa_data) if noaa_data else {}

    # Process AccuWeather data
    accu_daily = {}
    if accu_data:
        for d in accu_data:
            if 'high_f' in d and 'low_f' in d:
                accu_daily[d['date']] = {
                    'high_f': int(d['high_f']),
                    'low_f': int(d['low_f'])
                }
            else:
                accu_daily[d['date']] = {
                    'high_f': round(d['high_c'] * 1.8 + 32),
                    'low_f': round(d['low_c'] * 1.8 + 32)
                }

    # Process Google Weather data
    google_daily = {}
    if google_data:
        for d in google_data.get('daily', []):
            if 'high_f' in d and 'low_f' in d:
                google_daily[d['date']] = {
                    'high_f': int(d['high_f']),
                    'low_f': int(d['low_f'])
                }
            elif 'high_c' in d and 'low_c' in d:
                google_daily[d['date']] = {
                    'high_f': round(d['high_c'] * 1.8 + 32),
                    'low_f': round(d['low_c'] * 1.8 + 32)
                }

    # Process Weather.com data
    weather_com_daily = {}
    if weather_com_data:
        for d in weather_com_data:
            if 'high_f' in d and 'low_f' in d:
                weather_com_daily[d['date']] = {
                    'high_f': int(d['high_f']),
                    'low_f': int(d['low_f'])
                }

    # Process Weather Underground data
    wunderground_daily = {}
    if wunderground_data:
        for d in wunderground_data:
            if 'high_f' in d and 'low_f' in d:
                wunderground_daily[d['date']] = {
                    'high_f': int(d['high_f']),
                    'low_f': int(d['low_f'])
                }

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Forecast"

    # Use passed-in timestamp to ensure filename and content match
    if report_timestamp:
        report_time = report_timestamp
    else:
        report_time = datetime.now(ZoneInfo("America/Los_Angeles"))
    timestamp_str = report_time.strftime("%A, %B %d, %Y %H:%M:%S")

    # Styles
    thin_border = create_thin_border()
    header_font = Font(bold=True, size=14, color="003C78")
    title_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    small_font = Font(size=9)
    bold_small = Font(bold=True, size=9)
    center_align = Alignment(horizontal='center', vertical='center')

    current_row = 1

    # ===================
    # HEADER
    # ===================
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=18)
    header_cell = ws.cell(row=current_row, column=1, value='MODESTO, CA - DAILY WEATHER FORECAST')
    header_cell.font = header_font
    header_cell.alignment = center_align
    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=18)
    date_cell = ws.cell(row=current_row, column=1, value=timestamp_str)
    date_cell.font = title_font
    date_cell.alignment = center_align
    current_row += 2

    # ===================
    # MID WEATHER 48-HOUR SUMMARY (Top section)
    # ===================
    mid_start_col = 14
    mid_row = current_row

    ws.merge_cells(start_row=mid_row, start_column=mid_start_col, end_row=mid_row, end_column=mid_start_col + 4)
    mid_header = ws.cell(row=mid_row, column=mid_start_col, value='MID WEATHER 48-HOUR SUMMARY')
    mid_header.font = Font(bold=True, size=9, color="003C78")
    mid_header.alignment = center_align
    mid_header.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")

    if mid_data:
        today_data = mid_data.get('today', {})
        yest_data = mid_data.get('yesterday', {})

        # Headers row
        mid_row += 1
        for col_offset, label in enumerate(['', 'High', 'Low', 'Rain']):
            cell = ws.cell(row=mid_row, column=mid_start_col + col_offset, value=label)
            cell.font = bold_small
            cell.alignment = center_align
            cell.border = thin_border

        # TODAY row
        mid_row += 1
        ws.cell(row=mid_row, column=mid_start_col, value='TODAY').font = bold_small

        today_hi_cell = ws.cell(row=mid_row, column=mid_start_col + 1, value=f"{today_data.get('high', '--')}F")
        today_hi_cell.fill = PatternFill(start_color="FFC8B4", end_color="FFC8B4", fill_type="solid")
        today_hi_cell.alignment = center_align
        today_hi_cell.border = thin_border

        today_lo_cell = ws.cell(row=mid_row, column=mid_start_col + 2, value=f"{today_data.get('low', '--')}F")
        today_lo_cell.fill = PatternFill(start_color="B4D2FF", end_color="B4D2FF", fill_type="solid")
        today_lo_cell.alignment = center_align
        today_lo_cell.border = thin_border

        rain_cell = ws.cell(row=mid_row, column=mid_start_col + 3, value=f"{today_data.get('rain', '0.00')}\"")
        rain_cell.alignment = center_align
        rain_cell.border = thin_border

        # YESTERDAY row
        mid_row += 1
        ws.cell(row=mid_row, column=mid_start_col, value='YEST').font = bold_small

        yest_hi_cell = ws.cell(row=mid_row, column=mid_start_col + 1, value=f"{yest_data.get('high', '--')}F")
        yest_hi_cell.fill = PatternFill(start_color="FFC8B4", end_color="FFC8B4", fill_type="solid")
        yest_hi_cell.alignment = center_align
        yest_hi_cell.border = thin_border

        yest_lo_cell = ws.cell(row=mid_row, column=mid_start_col + 2, value=f"{yest_data.get('low', '--')}F")
        yest_lo_cell.fill = PatternFill(start_color="B4D2FF", end_color="B4D2FF", fill_type="solid")
        yest_lo_cell.alignment = center_align
        yest_lo_cell.border = thin_border

        yest_rain = ws.cell(row=mid_row, column=mid_start_col + 3, value=f"{yest_data.get('rain', '0.00')}\"")
        yest_rain.alignment = center_align
        yest_rain.border = thin_border

        # Historical records
        if 'record_high_temp' in mid_data:
            mid_row += 1
            rec_hi = mid_data.get('record_high_temp', '--')
            rec_hi_yr = mid_data.get('record_high_year', '')
            rec_lo = mid_data.get('record_low_temp', '--')
            rec_lo_yr = mid_data.get('record_low_year', '')
            ws.merge_cells(start_row=mid_row, start_column=mid_start_col, end_row=mid_row, end_column=mid_start_col + 3)
            rec_cell = ws.cell(row=mid_row, column=mid_start_col,
                             value=f'Records: Hi {rec_hi}F({rec_hi_yr}) Lo {rec_lo}F({rec_lo_yr})')
            rec_cell.font = Font(italic=True, size=8)

    current_row += 6

    # ===================
    # BUILD MERGED CONDITIONS MAP
    # ===================
    daily_conditions = {}

    # Step 1: Open-Meteo as base
    for day_record in om_daily:
        date_key = day_record.get('date', '')
        condition = day_record.get('condition', 'Unknown')
        if condition and condition != 'Unknown':
            daily_conditions[date_key] = {'condition': condition, 'source': 'Open-Meteo'}

    # Step 2: AccuWeather overwrites
    if accu_data:
        for day_record in accu_data:
            date_key = day_record.get('date', '')
            condition = day_record.get('condition', '')
            if condition and condition != 'Unknown':
                daily_conditions[date_key] = {'condition': condition, 'source': 'AccuWeather'}

    # Step 3: Google overwrites (best quality)
    if google_data:
        for day_record in google_data.get('daily', []):
            date_key = day_record.get('date', '')
            condition = day_record.get('condition', '')
            if condition and condition != 'Unknown':
                daily_conditions[date_key] = {'condition': condition, 'source': 'Google'}

    # Pre-calculate excluded highs (Open-Meteo max outlier exclusion)
    excluded_highs = {}
    weights = [1.0, 3.0, 3.0, 4.0, 4.0, 4.0, 6.0]

    for i, day in enumerate(om_daily):
        k = day.get('date', '')
        hi_vals = [
            day.get('high_f'),
            noaa_daily.get(k, {}).get('high_f'),
            met_daily.get(k, {}).get('high_f'),
            accu_daily.get(k, {}).get('high_f'),
            weather_com_daily.get(k, {}).get('high_f'),
            wunderground_daily.get(k, {}).get('high_f'),
            google_daily.get(k, {}).get('high_f')
        ]
        valid_highs = [(idx, v) for idx, v in enumerate(hi_vals) if v is not None]
        if valid_highs:
            max_high = max(v for _, v in valid_highs)
            om_high = hi_vals[0]
            if om_high is not None and om_high == max_high:
                excluded_highs[i] = {0}
            else:
                excluded_highs[i] = set()
        else:
            excluded_highs[i] = set()

    # ===================
    # CONDITION DESCRIPTORS ROW
    # ===================
    cond_row = current_row
    ws.cell(row=cond_row, column=1, value='').border = thin_border
    ws.cell(row=cond_row, column=2, value='').border = thin_border

    for i, day in enumerate(om_daily):
        date_key = day.get('date', '')
        condition_info = daily_conditions.get(date_key, {'condition': 'Unknown', 'source': None})
        condition = condition_info['condition']
        display_text, bg_color, text_color, is_special = get_daily_condition_display(condition)

        col = 3 + i * 2
        ws.merge_cells(start_row=cond_row, start_column=col, end_row=cond_row, end_column=col + 1)
        cell = ws.cell(row=cond_row, column=col, value=display_text)
        cell.fill = PatternFill(start_color=rgb_to_hex(*bg_color), end_color=rgb_to_hex(*bg_color), fill_type="solid")
        cell.font = Font(bold=is_special, size=9, color=rgb_to_hex(*text_color))
        cell.alignment = center_align
        cell.border = thin_border

    current_row += 1

    # ===================
    # TEMPERATURE GRID HEADER ROW (Day Names)
    # ===================
    header_fill = PatternFill(start_color="003C78", end_color="003C78", fill_type="solid")
    white_font = Font(bold=True, size=10, color="FFFFFF")

    # Weight column header
    ws.cell(row=current_row, column=1, value='').fill = header_fill
    ws.cell(row=current_row, column=1).border = thin_border

    # Source column header
    ws.cell(row=current_row, column=2, value='SOURCE').fill = header_fill
    ws.cell(row=current_row, column=2).font = white_font
    ws.cell(row=current_row, column=2).alignment = center_align
    ws.cell(row=current_row, column=2).border = thin_border

    # Day name headers
    for i, day in enumerate(om_daily):
        label = "TODAY" if i == 0 else day.get('day_name', '')[:3].upper()
        col = 3 + i * 2
        ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 1)

        base_color = DAY_COLORS[i % len(DAY_COLORS)]
        dark_color = (max(0, base_color[0] - 100), max(0, base_color[1] - 80), max(0, base_color[2] - 60))

        cell = ws.cell(row=current_row, column=col, value=label)
        cell.fill = PatternFill(start_color=rgb_to_hex(*dark_color), end_color=rgb_to_hex(*dark_color), fill_type="solid")
        cell.font = white_font
        cell.alignment = center_align
        cell.border = thin_border

    current_row += 1

    # ===================
    # DATES ROW
    # ===================
    ws.cell(row=current_row, column=1, value='').fill = PatternFill(start_color="466EA0", end_color="466EA0", fill_type="solid")
    ws.cell(row=current_row, column=1).border = thin_border

    date_header = ws.cell(row=current_row, column=2, value='DATE')
    date_header.fill = PatternFill(start_color="466EA0", end_color="466EA0", fill_type="solid")
    date_header.font = Font(size=9, color="FFFFFF")
    date_header.alignment = center_align
    date_header.border = thin_border

    for i, day in enumerate(om_daily):
        date_str = day.get('date', '')[5:]  # MM-DD
        col = 3 + i * 2
        ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 1)

        base_color = DAY_COLORS[i % len(DAY_COLORS)]
        dark_color = (max(0, base_color[0] - 70), max(0, base_color[1] - 50), max(0, base_color[2] - 30))

        cell = ws.cell(row=current_row, column=col, value=date_str)
        cell.fill = PatternFill(start_color=rgb_to_hex(*dark_color), end_color=rgb_to_hex(*dark_color), fill_type="solid")
        cell.font = Font(size=9, color="FFFFFF")
        cell.alignment = center_align
        cell.border = thin_border

    current_row += 1

    # ===================
    # SOURCE ROWS
    # ===================
    def write_source_row(row_num: int, label: str, getter, source_idx: int):
        """Write a single source row with Hi/Lo values."""
        # Weight column
        weight_val = SOURCE_WEIGHT_DISPLAY.get(label, '')
        weight_cell = ws.cell(row=row_num, column=1, value=weight_val)
        weight_cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        weight_cell.font = small_font
        weight_cell.alignment = center_align
        weight_cell.border = thin_border

        # Source label - with clickable hyperlink if URL exists
        source_cell = ws.cell(row=row_num, column=2, value=label)
        source_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        source_cell.alignment = center_align
        source_cell.border = thin_border
        if label in SOURCE_URLS:
            source_cell.hyperlink = SOURCE_URLS[label]
            source_cell.font = Font(size=7, bold=True, color='0000FF', underline='single')
        else:
            source_cell.font = bold_small

        # Temperature values
        for i, d in enumerate(om_daily):
            v1, v2 = getter(d, d.get('date', ''))
            is_excluded = source_idx in excluded_highs.get(i, set())
            day_color = DAY_COLORS[i % len(DAY_COLORS)]
            fill = PatternFill(start_color=rgb_to_hex(*day_color), end_color=rgb_to_hex(*day_color), fill_type="solid")

            # High cell
            hi_col = 3 + i * 2
            hi_val = "-" if (is_excluded and v1 is not None) else (str(v1) if v1 else "--")
            hi_cell = ws.cell(row=row_num, column=hi_col, value=hi_val)
            hi_cell.fill = fill
            hi_cell.font = normal_font
            hi_cell.alignment = center_align
            hi_cell.border = thin_border

            # Low cell
            lo_col = hi_col + 1
            lo_cell = ws.cell(row=row_num, column=lo_col, value=str(v2) if v2 else "--")
            lo_cell.fill = fill
            lo_cell.font = normal_font
            lo_cell.alignment = center_align
            lo_cell.border = thin_border

    # Write all source rows
    write_source_row(current_row, 'OPEN-METEO',
                    lambda d, k: (d.get('high_f'), d.get('low_f')), 0)
    current_row += 1

    write_source_row(current_row, 'NOAA (GOV)',
                    lambda d, k: (noaa_daily.get(k, {}).get('high_f'), noaa_daily.get(k, {}).get('low_f')), 1)
    current_row += 1

    write_source_row(current_row, 'MET.NO (EU)',
                    lambda d, k: (met_daily.get(k, {}).get('high_f'), met_daily.get(k, {}).get('low_f')), 2)
    current_row += 1

    write_source_row(current_row, 'ACCUWEATHER',
                    lambda d, k: (accu_daily.get(k, {}).get('high_f'), accu_daily.get(k, {}).get('low_f')), 3)
    current_row += 1

    write_source_row(current_row, 'WEATHER.COM',
                    lambda d, k: (weather_com_daily.get(k, {}).get('high_f'), weather_com_daily.get(k, {}).get('low_f')), 4)
    current_row += 1

    write_source_row(current_row, 'WUNDERGRND',
                    lambda d, k: (wunderground_daily.get(k, {}).get('high_f'), wunderground_daily.get(k, {}).get('low_f')), 5)
    current_row += 1

    write_source_row(current_row, 'GOOGLE (AI)',
                    lambda d, k: (google_daily.get(k, {}).get('high_f'), google_daily.get(k, {}).get('low_f')), 6)
    current_row += 1

    # ===================
    # WEIGHTED AVERAGES ROW
    # ===================
    avg_fill = PatternFill(start_color="FFDC64", end_color="FFDC64", fill_type="solid")

    ws.cell(row=current_row, column=1, value='').fill = avg_fill
    ws.cell(row=current_row, column=1).border = thin_border

    avg_label = ws.cell(row=current_row, column=2, value='Wtd. Averages')
    avg_label.fill = avg_fill
    avg_label.font = bold_small
    avg_label.alignment = center_align
    avg_label.border = thin_border

    for i, day in enumerate(om_daily):
        k = day.get('date', '')
        day_color = DAY_COLORS[i % len(DAY_COLORS)]
        avg_color = (min(255, day_color[0] + 10), min(255, day_color[1] - 10), max(0, day_color[2] - 40))
        fill = PatternFill(start_color=rgb_to_hex(*avg_color), end_color=rgb_to_hex(*avg_color), fill_type="solid")

        hi_vals = [
            day.get('high_f'),
            noaa_daily.get(k, {}).get('high_f'),
            met_daily.get(k, {}).get('high_f'),
            accu_daily.get(k, {}).get('high_f'),
            weather_com_daily.get(k, {}).get('high_f'),
            wunderground_daily.get(k, {}).get('high_f'),
            google_daily.get(k, {}).get('high_f')
        ]
        lo_vals = [
            day.get('low_f'),
            noaa_daily.get(k, {}).get('low_f'),
            met_daily.get(k, {}).get('low_f'),
            accu_daily.get(k, {}).get('low_f'),
            weather_com_daily.get(k, {}).get('low_f'),
            wunderground_daily.get(k, {}).get('low_f'),
            google_daily.get(k, {}).get('low_f')
        ]

        avg_hi, _ = calculate_weighted_average_excluding_om_max(hi_vals, weights)
        avg_lo = calculate_weighted_average(lo_vals, weights)

        hi_col = 3 + i * 2
        hi_cell = ws.cell(row=current_row, column=hi_col, value=str(avg_hi) if avg_hi else "--")
        hi_cell.fill = fill
        hi_cell.font = Font(bold=True, size=10)
        hi_cell.alignment = center_align
        hi_cell.border = thin_border

        lo_cell = ws.cell(row=current_row, column=hi_col + 1, value=str(avg_lo) if avg_lo else "--")
        lo_cell.fill = fill
        lo_cell.font = Font(bold=True, size=10)
        lo_cell.alignment = center_align
        lo_cell.border = thin_border

    current_row += 1

    # ===================
    # PRECIPITATION ROW
    # ===================
    precip_fill = PatternFill(start_color="B4D2FF", end_color="B4D2FF", fill_type="solid")

    ws.cell(row=current_row, column=1, value='').fill = precip_fill
    ws.cell(row=current_row, column=1).border = thin_border

    precip_label = ws.cell(row=current_row, column=2, value='PRECIP %')
    precip_label.fill = precip_fill
    precip_label.font = bold_small
    precip_label.alignment = center_align
    precip_label.border = thin_border

    for i, day in enumerate(om_daily):
        k = day.get('date', '')
        precip_pct = 0
        if precip_data and k in precip_data:
            precip_pct = precip_data[k].get('consensus', 0)
        else:
            precip_pct = day.get('precip_prob', 0)

        # Color based on probability
        if precip_pct >= 50:
            fill = PatternFill(start_color="6496FF", end_color="6496FF", fill_type="solid")
        elif precip_pct >= 25:
            fill = PatternFill(start_color="B4D2FF", end_color="B4D2FF", fill_type="solid")
        else:
            day_color = DAY_COLORS[i % len(DAY_COLORS)]
            fill = PatternFill(start_color=rgb_to_hex(*day_color), end_color=rgb_to_hex(*day_color), fill_type="solid")

        col = 3 + i * 2
        ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 1)
        cell = ws.cell(row=current_row, column=col, value=f"{precip_pct}%")
        cell.fill = fill
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border

    current_row += 1

    # Precip sources note
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=18)
    note = ws.cell(row=current_row, column=1,
                  value='PRECIP = Google (0-72hr) > AccuWeather (72hr+) > Open-Meteo')
    note.font = Font(italic=True, size=8, color="505050")
    note.alignment = Alignment(horizontal='right')
    current_row += 2

    # ===================
    # SOLAR FORECAST GRID
    # ===================
    solar_header = ws.cell(row=current_row, column=1,
                          value='SOLAR FORECAST (GOOGLE AI WEATHER API) - W/m² Irradiance')
    solar_header.font = Font(bold=True, size=11, color="003C78")
    current_row += 1

    tz = ZoneInfo("America/Los_Angeles")
    forecast_dates = [(datetime.now(tz) + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(0, 4)]

    # Build duck curve data from Google Weather
    duck_data = {d: [] for d in forecast_dates}
    google_hourly = google_data.get('hourly', []) if google_data else []

    for hour_record in google_hourly:
        try:
            time_str = hour_record.get('time', '')
            if not time_str:
                continue

            if 'Z' in time_str:
                dt = datetime.fromisoformat(time_str.replace('Z', '+00:00')).astimezone(tz)
            else:
                dt = datetime.fromisoformat(time_str).astimezone(tz)

            row_date = dt.strftime('%Y-%m-%d')
            row_hour = dt.hour

            if row_date in forecast_dates and 9 <= row_hour <= 16:
                cloud_cover = hour_record.get('cloud_cover', 50)
                day_of_year = dt.timetuple().tm_yday
                condition = hour_record.get('condition', 'Unknown')
                irradiance = estimate_irradiance_from_cloud_cover(cloud_cover, row_hour, day_of_year)

                condition_lower = condition.lower()
                if 'rain' in condition_lower or 'storm' in condition_lower:
                    risk = 'HIGH'
                elif cloud_cover >= 90:
                    risk = 'MODERATE'
                elif cloud_cover >= 70:
                    risk = 'LOW-MOD'
                else:
                    risk = 'LOW'

                duck_data[row_date].append({
                    'hour': row_hour,
                    'solar': irradiance,
                    'risk': risk,
                    'condition': condition
                })
        except Exception:
            continue

    # Fallback to df_analyzed if no Google data
    if not any(duck_data.values()) and df_analyzed is not None:
        for _, row in df_analyzed.iterrows():
            try:
                row_date = row['time'].strftime('%Y-%m-%d')
                row_hour = row['time'].hour
                if row_date in forecast_dates and 9 <= row_hour <= 16:
                    duck_data[row_date].append({
                        'hour': row_hour,
                        'solar': row.get('solar_adjusted', 0),
                        'risk': row.get('risk_level', 'LOW'),
                        'condition': None
                    })
            except Exception:
                continue

    # Solar header row
    solar_header_fill = PatternFill(start_color="003C78", end_color="003C78", fill_type="solid")

    ws.cell(row=current_row, column=1, value='DATE').fill = solar_header_fill
    ws.cell(row=current_row, column=1).font = white_font
    ws.cell(row=current_row, column=1).alignment = center_align
    ws.cell(row=current_row, column=1).border = thin_border

    hour_labels = ['9AM', '10', '11', '12PM', '1', '2', '3', '4PM']
    for i, hl in enumerate(hour_labels):
        col = 2 + i
        cell = ws.cell(row=current_row, column=col, value=hl)
        cell.fill = solar_header_fill
        cell.font = white_font
        cell.alignment = center_align
        cell.border = thin_border

    current_row += 1

    # Solar data rows (2 rows per day: values + descriptions)
    for d in forecast_dates:
        date_obj = datetime.strptime(d, '%Y-%m-%d')
        day_name = date_obj.strftime('%A')
        date_label = f"{d[5:]} {day_name}"

        # Merge date cell across 2 rows
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=1)
        date_cell = ws.cell(row=current_row, column=1, value=date_label)
        date_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        date_cell.font = bold_small
        date_cell.alignment = center_align
        date_cell.border = thin_border

        hours_dict = {h['hour']: h for h in duck_data.get(d, [])}

        # Values row
        for i in range(8):
            h_data = hours_dict.get(9 + i, {'solar': 0, 'risk': 'LOW', 'condition': 'Unknown'})
            solar_display = h_data['solar'] * 1.15  # 15% boost
            condition = h_data.get('condition', 'Unknown')

            (r, g, b), risk_desc = get_solar_color_and_desc(h_data['risk'], solar_display, condition)
            fill = PatternFill(start_color=rgb_to_hex(r, g, b), end_color=rgb_to_hex(r, g, b), fill_type="solid")

            col = 2 + i
            val_cell = ws.cell(row=current_row, column=col, value=f"{solar_display:.0f}")
            val_cell.fill = fill
            val_cell.font = normal_font
            val_cell.alignment = center_align
            val_cell.border = thin_border

        current_row += 1

        # Description row
        for i in range(8):
            h_data = hours_dict.get(9 + i, {'solar': 0, 'risk': 'LOW', 'condition': 'Unknown'})
            solar_display = h_data['solar'] * 1.15
            condition = h_data.get('condition', 'Unknown')

            (r, g, b), risk_desc = get_solar_color_and_desc(h_data['risk'], solar_display, condition)
            fill = PatternFill(start_color=rgb_to_hex(r, g, b), end_color=rgb_to_hex(r, g, b), fill_type="solid")

            col = 2 + i
            desc_cell = ws.cell(row=current_row, column=col, value=risk_desc)
            desc_cell.fill = fill
            desc_cell.font = Font(italic=True, size=8)
            desc_cell.alignment = center_align
            desc_cell.border = thin_border

        current_row += 1

    # ===================
    # SOLAR LEGEND
    # ===================
    current_row += 1
    legend_items = [
        ('DCDCDC', 'Cloudy'),
        ('C8E6FF', 'Some Sun'),
        ('C8FFC8', 'Good Sun'),
        ('90EE90', 'Full Sun'),
        ('FFFFB4', 'Fog Possible'),
        ('FFD2A0', 'Heavy Clouds'),
        ('FFB4B4', 'Dense Fog'),
        ('B4A0C8', 'Tule Fog'),
    ]

    ws.cell(row=current_row, column=1, value='Legend:').font = Font(bold=True, size=8)
    col = 2
    for color, label in legend_items:
        cell = ws.cell(row=current_row, column=col, value=label)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(size=8)
        cell.alignment = center_align
        cell.border = thin_border
        col += 1

    ws.cell(row=current_row, column=col, value='(values = W/m²)').font = Font(italic=True, size=7)

    # ===================
    # ADJUST COLUMN WIDTHS
    # ===================
    ws.column_dimensions['A'].width = 8   # Weight column
    ws.column_dimensions['B'].width = 14  # Source column
    for i in range(3, 20):
        ws.column_dimensions[get_column_letter(i)].width = 7

    # ===================
    # SAVE WORKBOOK
    # ===================
    if output_path is None:
        pacific = ZoneInfo("America/Los_Angeles")
        now = datetime.now(pacific)
        timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
        output_path = Path("reports") / now.strftime("%Y-%m") / now.strftime("%Y-%m-%d") / f"daily_forecast_{timestamp}.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb.save(str(output_path))
        logger.info(f"[generate_xlsx_report] Excel saved to: {output_path}")
        return output_path
    except Exception as e:
        logger.error(f"[generate_xlsx_report] Failed to save Excel: {e}", exc_info=True)
        return None


if __name__ == "__main__":
    import asyncio
    from duck_sun.providers.open_meteo import fetch_open_meteo
    from duck_sun.providers.noaa import NOAAProvider
    from duck_sun.providers.met_no import MetNoProvider
    from duck_sun.providers.accuweather import AccuWeatherProvider
    from duck_sun.uncanniness import UncannyEngine
    from dotenv import load_dotenv

    load_dotenv()
    logging.basicConfig(level=logging.INFO)

    async def test():
        print("=== Testing Excel Report Generator ===\n")

        om_data = await fetch_open_meteo(days=8)

        noaa = NOAAProvider()
        noaa_data = await noaa.fetch_async()

        met = MetNoProvider()
        met_data = await met.fetch_async()

        accu = AccuWeatherProvider()
        accu_data = await accu.fetch_forecast()

        engine = UncannyEngine()
        df = engine.normalize_temps(om_data, noaa_data, met_data)
        df_analyzed = engine.analyze_duck_curve(df)

        critical = len(df_analyzed[df_analyzed['risk_level'].str.contains('CRITICAL', na=False)])

        xlsx_path = generate_xlsx_report(
            om_data=om_data,
            noaa_data=noaa_data,
            met_data=met_data,
            accu_data=accu_data,
            df_analyzed=df_analyzed,
            fog_critical_hours=critical
        )

        if xlsx_path:
            print(f"\n Excel generated: {xlsx_path}")

    asyncio.run(test())
